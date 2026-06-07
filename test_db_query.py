import openpyxl
import os

# Check what data is in the "4. Budget Control" sheet
excel_path = 'plans/DirOps/1LC-522UD019 - 13 UNIT RADAR GCI DAN PENDUKUNGNYA.xlsx'

if os.path.exists(excel_path):
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    
    # Read the Budget Control sheet
    if '4. Budget Control' in wb.sheetnames:
        ws = wb['4. Budget Control']
        print("First 20 rows of '4. Budget Control' sheet:")
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= 20:
                break
            print(f"Row {i+1}: {row[:8]}...")  # First 8 columns
    else:
        print("Sheet '4. Budget Control' not found")
    
    # Also check Project Profile for contract value
    if '1. Project Profile' in wb.sheetnames:
        ws = wb['1. Project Profile']
        print("\n\nFirst 20 rows of '1. Project Profile' sheet:")
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= 20:
                break
            print(f"Row {i+1}: {row[:4]}...")
    
    wb.close()
else:
    print(f"File not found: {excel_path}")

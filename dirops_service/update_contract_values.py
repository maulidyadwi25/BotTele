"""Update existing project records with correct contract values from Excel files.

This script re-parses the Excel files and updates the contract_value_idr,
contract_value_valas, and currency fields for existing projects.
"""
import os
import sys
from pathlib import Path
from decimal import Decimal, InvalidOperation

import openpyxl

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from dirops_service.database import create_app
from dirops_service.models import Project


def parse_value(val):
    """Parse cell value to appropriate Python type."""
    if val is None or val == '':
        return None
    if isinstance(val, (int, float)):
        return val
    val = str(val).strip()

    if val == '':
        return None

    # Try parsing as number
    for typ in (float, int):
        try:
            cleaned = val.replace('%', '').replace(' ', '')
            if ',' in cleaned:
                parts = cleaned.split(',')
                if len(parts) == 2 and len(parts[1]) <= 2:
                    cleaned = parts[0].replace('.', '') + '.' + parts[1]
                else:
                    cleaned = cleaned.replace(',', '')
            elif '.' in cleaned:
                dot_count = cleaned.count('.')
                if dot_count >= 2:
                    cleaned = cleaned.replace('.', '')
            return typ(cleaned)
        except (ValueError, InvalidOperation):
            pass

    return val


def parse_project_profile(ws):
    """Parse Project Profile sheet."""
    data = {}
    for row in range(1, ws.max_row + 1):
        key_cell = ws.cell(row=row, column=2).value
        val_cell = ws.cell(row=row, column=3).value
        unit_cell = ws.cell(row=row, column=4).value
        
        if key_cell:
            key = str(key_cell).strip()
            
            if 'Nilai Proyek (IDR)' in key and val_cell is not None:
                data['Nilai Proyek (IDR)'] = parse_value(val_cell)
                if unit_cell:
                    currency = str(unit_cell).strip()
                    if currency and currency != 'None':
                        data['Currency IDR'] = currency
            elif 'Nilai Proyek (Valas)' in key and val_cell is not None:
                data['Nilai Proyek (Valas)'] = parse_value(val_cell)
                if unit_cell:
                    currency = str(unit_cell).strip()
                    if currency and currency != 'None':
                        data['Currency Valas'] = currency
            elif val_cell is not None:
                data[key] = parse_value(val_cell)
    return data


def update_project_from_excel(project_code, excel_path):
    """Update a single project record from Excel file."""
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        
        profile = {}
        if '1. Project Profile' in wb.sheetnames:
            ws = wb['1. Project Profile']
            profile = parse_project_profile(ws)
        
        wb.close()
        
        return {
            'contract_value_idr': profile.get('Nilai Proyek (IDR)'),
            'contract_value_valas': profile.get('Nilai Proyek (Valas)'),
            'currency': profile.get('Currency Valas') or profile.get('Currency IDR'),
        }
    except Exception as e:
        print(f"Error parsing {excel_path}: {e}")
        return None


def main():
    DIROPS_FOLDER = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'plans', 'DirOps')
    
    app = create_app()
    updated_count = 0
    
    with app.app_context():
        projects = Project.query.all()
        print(f"Found {len(projects)} projects in database")
        
        for project in projects:
            # Find the Excel file for this project
            excel_path = None
            for xlsx_file in Path(DIROPS_FOLDER).glob('*.xlsx'):
                code = xlsx_file.stem.split(' - ')[0]
                if code == project.project_code:
                    excel_path = str(xlsx_file)
                    break
            
            if not excel_path:
                print(f"  [{project.project_code}] Excel file not found, skipping...")
                continue
            
            print(f"  [{project.project_code}] Updating from {Path(excel_path).name}...")
            
            new_values = update_project_from_excel(project.project_code, excel_path)
            if not new_values:
                continue
            
            old_idr = float(project.contract_value_idr) if project.contract_value_idr else None
            old_valas = float(project.contract_value_valas) if project.contract_value_valas else None
            old_currency = project.currency
            
            new_idr = new_values['contract_value_idr']
            new_valas = new_values['contract_value_valas']
            new_currency = new_values['currency']
            
            # Check if values are different
            changed = False
            if old_idr != new_idr:
                print(f"    contract_value_idr: {old_idr} -> {new_idr}")
                project.contract_value_idr = new_idr
                changed = True
            if old_valas != new_valas:
                print(f"    contract_value_valas: {old_valas} -> {new_valas}")
                project.contract_value_valas = new_valas
                changed = True
            if old_currency != new_currency:
                print(f"    currency: {old_currency} -> {new_currency}")
                project.currency = new_currency
                changed = True
            
            if changed:
                updated_count += 1
        
        if updated_count > 0:
            print(f"\nCommitting {updated_count} changes...")
            from dirops_service.models import db
            db.session.commit()
            print("Done!")
        else:
            print("\nNo changes needed.")


if __name__ == '__main__':
    main()
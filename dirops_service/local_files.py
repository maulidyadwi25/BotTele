"""Local Excel File Service for DirOps

Reads project data from local Excel files in plans/DirOps instead of Google Sheets.
"""
import os
import sys
from pathlib import Path
from typing import List, Dict, Optional, Tuple
import openpyxl

# Ensure we're using the local dirops.db
DIROPS_DB_PATH = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'dirops.db')
DIROPS_FOLDER = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'plans', 'DirOps')

# Cache for Excel file data
EXCEL_CACHE = {}
EXCEL_CACHE_TTL = 300  # 5 minutes


def get_project_excel_path(project_code: str) -> Optional[str]:
    """Get the Excel file path for a project code."""
    folder = Path(DIROPS_FOLDER)
    if not folder.exists():
        return None
    
    for xlsx_file in folder.glob('*.xlsx'):
        code = xlsx_file.stem.split(' - ')[0]
        if code == project_code:
            return str(xlsx_file)
    return None


def get_project_excel_name(project_code: str) -> Optional[str]:
    """Get the Excel file name for a project code."""
    path = get_project_excel_path(project_code)
    if path:
        return Path(path).name
    return None


def get_project_sheets(project_code: str) -> List[str]:
    """Get list of sheet names from a project's Excel file."""
    xlsx_path = get_project_excel_path(project_code)
    if not xlsx_path:
        return []
    
    try:
        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        sheets = wb.sheetnames
        wb.close()
        return sheets
    except Exception as e:
        print(f"Error reading sheets from {xlsx_path}: {e}")
        return []


def get_project_sheet_data(project_code: str, sheet_name: str) -> List[List]:
    """Get data from a specific sheet in a project's Excel file."""
    xlsx_path = get_project_excel_path(project_code)
    if not xlsx_path:
        return []
    
    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        if sheet_name not in wb.sheetnames:
            wb.close()
            return []
        
        ws = wb[sheet_name]
        data = []
        
        for row in ws.iter_rows(values_only=True):
            # Convert row to list, handling various types
            row_data = []
            for cell in row:
                if cell is None:
                    row_data.append(None)
                elif isinstance(cell, (int, float, str)):
                    row_data.append(cell)
                else:
                    row_data.append(str(cell))
            data.append(row_data)
        
        wb.close()
        return data
    except Exception as e:
        print(f"Error reading {sheet_name} from {xlsx_path}: {e}")
        return []


def get_project_profile(project_code: str) -> Dict:
    """Get project profile data from the Project Profile sheet."""
    xlsx_path = get_project_excel_path(project_code)
    if not xlsx_path:
        return {}
    
    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        if '1. Project Profile' not in wb.sheetnames:
            wb.close()
            return {}
        
        ws = wb['1. Project Profile']
        profile = {}
        
        for row in ws.iter_rows(values_only=True):
            # In "1. Project Profile" sheet:
            # Column A (index 0) is empty, Column B (index 1) is key, Column C (index 2) is value, Column D (index 3) is unit
            if len(row) >= 3 and row[1]:
                key = str(row[1]).strip()
                value = row[2]
                if value is not None:
                    if isinstance(value, (int, float)):
                        profile[key] = value
                    else:
                        profile[key] = str(value).strip()
                # Also store unit/currency if present
                if len(row) >= 4 and row[3]:
                    unit = str(row[3]).strip()
                    if unit and unit != 'None':
                        profile[f'{key} Unit'] = unit
        
        wb.close()
        return profile
    except Exception as e:
        print(f"Error reading profile from {xlsx_path}: {e}")
        return {}


def get_all_project_codes() -> List[str]:
    """Get all project codes from local Excel files."""
    folder = Path(DIROPS_FOLDER)
    if not folder.exists():
        return []
    
    codes = []
    for xlsx_file in folder.glob('*.xlsx'):
        code = xlsx_file.stem.split(' - ')[0]
        if code not in codes:
            codes.append(code)
    
    return sorted(codes)


def get_all_projects_from_files() -> List[Dict]:
    """Get all project info from local Excel files."""
    folder = Path(DIROPS_FOLDER)
    if not folder.exists():
        return []
    
    projects = []
    for xlsx_file in folder.glob('*.xlsx'):
        stem = xlsx_file.stem
        parts = stem.split(' - ')
        code = parts[0]
        name = ' - '.join(parts[1:]) if len(parts) > 1 else stem
        
        projects.append({
            'code': code,
            'name': name,
            'file_name': xlsx_file.name,
            'file_path': str(xlsx_file)
        })
    
    return sorted(projects, key=lambda x: x['code'])


def get_dirops_files() -> List[Dict]:
    """Get list of DirOps files (for compatibility with bot.py gsheets interface)."""
    projects = get_all_projects_from_files()
    return [
        {
            'id': p['code'],  # Use project code as ID
            'name': p['file_name'],
            'code': p['code'],
            'project_name': p['name']
        }
        for p in projects
    ]


def get_dirops_sheets(project_code: str) -> List[str]:
    """Get list of sheets for a project (for compatibility with bot.py gsheets interface)."""
    return get_project_sheets(project_code)


def get_dirops_sheet_data(project_code: str, sheet_name: str) -> List[List]:
    """Get sheet data for a project (for compatibility with bot.py gsheets interface)."""
    return get_project_sheet_data(project_code, sheet_name)


def get_files_in_folder(folder_id: str = None) -> List[Dict]:
    """Get files in a folder - for local files, returns all files (no folder hierarchy)."""
    # For DirOps local files, we don't have folder hierarchy
    # Return all files as if they're in the root folder
    return get_dirops_files()


def get_all_data_from_files() -> List[Dict]:
    """Get all data from all DirOps files - for compatibility with bot.py."""
    files = get_dirops_files()
    all_data = []
    
    for f in files:
        project_code = f['code']
        sheets = get_dirops_sheets(project_code)
        
        for sheet_name in sheets:
            data = get_dirops_sheet_data(project_code, sheet_name)
            if data:
                all_data.append({
                    'file_id': project_code,
                    'file_name': f['name'],
                    'sheet_name': sheet_name,
                    'data': data
                })
    
    return all_data


if __name__ == '__main__':
    # Test
    print("Project codes found:")
    for code in get_all_project_codes():
        print(f"  {code}")
    
    print("\nFiles:")
    for f in get_dirops_files()[:3]:
        print(f"  {f['id']}: {f['name']}")
        print(f"    Sheets: {get_dirops_sheets(f['id'])[:5]}")
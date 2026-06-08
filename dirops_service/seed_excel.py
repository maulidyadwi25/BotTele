"""Excel to Database Seeder

Reads Kertas Kerja Monitoring Proyek Excel files and seeds them into the database.
"""
import os
import sys
import json
from datetime import datetime, date
from decimal import Decimal, InvalidOperation
from pathlib import Path

import openpyxl

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from dirops_service.database import create_app
from dirops_service.models import (
    db, Project, ProjectMilestone, WbsItem, WeeklyProgress,
    RecoveryPlan, ActionTracker, BudgetControl, BudgetDetailLine,
    CashflowIn, CashflowOut, Vendor, ProcurementItem, ProcurementPayment,
    ProcOpsStatus, ProcPolStatus, ProjectPhoto, ExchangeRate
)


def parse_value(val):
    """Parse cell value to appropriate Python type."""
    if val is None or val == '':
        return None
    if isinstance(val, (int, float)):
        return val
    if isinstance(val, datetime):
        return val
    if isinstance(val, date):
        return val
    val = str(val).strip()

    if val == '':
        return None

    # Try parsing as number
    # Handle Indonesian number format: 2.735.778.108 (dots as thousand sep) or 2.735.778.108,50 (decimal comma)
    for typ in (float, int):
        try:
            cleaned = val.replace('%', '').replace(' ', '')
            # Check if it's Indonesian format (dots as thousand separators)
            # Pattern: dots appear between groups of 3 digits
            if ',' in cleaned:
                # Has comma - could be decimal separator (Indonesian) or thousand separator (US)
                parts = cleaned.split(',')
                if len(parts) == 2 and len(parts[1]) <= 2:
                    # Likely decimal comma (Indonesian): 2.735.778.108,50
                    cleaned = parts[0].replace('.', '') + '.' + parts[1]
                else:
                    # Likely thousand separator comma: 2,735,778.108
                    cleaned = cleaned.replace(',', '')
            elif '.' in cleaned:
                # Check if dots are thousand separators (Indonesian)
                # If we have groups of 3 digits separated by dots, it's Indonesian format
                dot_count = cleaned.count('.')
                if dot_count >= 2:
                    # Likely Indonesian: 2.735.778.108
                    cleaned = cleaned.replace('.', '')
            return typ(cleaned)
        except (ValueError, InvalidOperation):
            pass

    # Try parsing date
    date_formats = ['%m/%d/%Y', '%d/%m/%Y', '%Y-%m-%d', '%d-%b-%Y', '%d-%m-%Y']
    for fmt in date_formats:
        try:
            return datetime.strptime(val, fmt).date()
        except ValueError:
            pass

    return val


def get_cell_value(ws, row, col):
    """Get cell value with parsing."""
    try:
        return parse_value(ws.cell(row=row, column=col).value)
    except:
        return None


def parse_project_profile(ws):
    """Parse Project Profile sheet."""
    data = {}
    for row in range(1, ws.max_row + 1):
        # Key is in column 2 (B), value is in column 3 (C), currency/unit in column 4 (D)
        key_cell = ws.cell(row=row, column=2).value
        val_cell = ws.cell(row=row, column=3).value
        unit_cell = ws.cell(row=row, column=4).value
        
        if key_cell:
            key = str(key_cell).strip()
            
            # Handle "Nilai Proyek (IDR)" and "Nilai Proyek (Valas)" specially
            # They have the currency/unit in column 4 (D)
            if 'Nilai Proyek (IDR)' in key and val_cell is not None:
                data['Nilai Proyek (IDR)'] = parse_value(val_cell)
                if unit_cell:
                    currency = str(unit_cell).strip()
                    if currency and currency != 'None':
                        data['Currency IDR'] = currency
            elif 'Nilai Proyek (Valas)' in key and val_cell is not None:
                data['Nilai Proyek (Valas)'] = parse_value(val_cell)
                if unit_cell:
                    currency = str(unit_cell).strip()
                    if currency and currency != 'None':
                        data['Currency Valas'] = currency
            elif val_cell is not None:
                data[key] = parse_value(val_cell)
    return data


def seed_from_excel(excel_path):
    """Seed database from a single Excel file."""
    print(f"Processing: {excel_path}")

    wb = openpyxl.load_workbook(excel_path, data_only=True)
    file_name = Path(excel_path).stem

    # Extract project code from filename (before first ' - ')
    project_code = file_name.split(' - ')[0] if ' - ' in file_name else file_name
    project_name = ' - '.join(file_name.split(' - ')[1:]) if ' - ' in file_name else file_name

    app = create_app()
    with app.app_context():
        # Check if project exists
        existing = Project.query.filter_by(project_code=project_code).first()
        if existing:
            print(f"  Project {project_code} already exists, skipping...")
            return

        # Parse Project Profile
        profile = {}
        if '1. Project Profile' in wb.sheetnames:
            ws = wb['1. Project Profile']
            profile = parse_project_profile(ws)

        # Create Project
        project = Project(
            project_code=project_code,
            project_name=project_name,
            customer=profile.get('Customer'),
            business_scheme=profile.get('Skema Bisnis'),
            status=profile.get('Status Proyek'),
            unit_kerja=profile.get('Unit Kerja'),
            contract_value_idr=profile.get('Nilai Proyek (IDR)'),
            contract_value_valas=profile.get('Nilai Proyek (Valas)'),
            currency=profile.get('Currency Valas') or profile.get('Currency IDR'),
            cogs_percent=profile.get('COGS %'),
            cogs_idr=profile.get('COGS IDR'),
            gpm_percent=profile.get('GPM %'),
            gpm_idr=profile.get('GPM IDR'),
            project_leader=profile.get('Pimpinan Proyek'),
            project_analyst=profile.get('Analis Proyek'),
            lkp_des_2025=profile.get('% LKP Des 2025'),
            target_rkap_2026=profile.get('Target RKAP 2026'),
            target_progress_2026=profile.get('Target +Progress 2026'),
            progress_achieved=profile.get('% Tercapai'),
            accumulated_progress=profile.get('% Akumulasi Progress'),
        )

        db.session.add(project)
        db.session.flush()

        print(f"  Created project: {project_code}")

        # Parse and seed other sheets...
        # (Full implementation would parse each sheet)

        db.session.commit()
        print(f"  Committed: {project_code}")


def seed_from_folder(folder_path):
    """Seed database from all Excel files in folder."""
    folder = Path(folder_path)
    # Skip temporary files (~$...)
    excel_files = [f for f in folder.glob('*.xlsx') if not f.name.startswith('~$')]
    excel_files += [f for f in folder.glob('*.xls') if not f.name.startswith('~$')]

    if not excel_files:
        print(f"No Excel files found in {folder_path}")
        return

    print(f"Found {len(excel_files)} Excel files")

    for excel_file in excel_files:
        try:
            seed_from_excel(excel_file)
        except Exception as e:
            print(f"  Error processing {excel_file.name}: {e}")
            import traceback
            traceback.print_exc()


if __name__ == '__main__':
    import argparse
    parser = argparse.ArgumentParser(description='Seed database from Excel files')
    parser.add_argument('path', help='Path to Excel file or folder containing Excel files')
    args = parser.parse_args()

    path = Path(args.path)
    if path.is_file():
        seed_from_excel(path)
    elif path.is_dir():
        seed_from_folder(path)
    else:
        print(f"Path not found: {path}")
        sys.exit(1)

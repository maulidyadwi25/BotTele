"""
Download spreadsheets from Google Drive folder as flat values (no formulas).
Downloads shortcut targets as actual spreadsheet data.
"""
import os
import re
from googleapiclient.discovery import build
from google.oauth2.service_account import Credentials
from googleapiclient.http import MediaIoBaseDownload
import io
import openpyxl

SERVICE_ACCOUNT_FILE = 'credentials.json'
FOLDER_ID = '1F0V5veGi6DH-uetS2V9UpBnlph7k8Q5w'
OUTPUT_FOLDER = 'plans/DirOps'

SCOPES = ['https://www.googleapis.com/auth/drive.readonly']


def get_drive_service():
    """Create and return Google Drive API service."""
    creds = Credentials.from_service_account_file(SERVICE_ACCOUNT_FILE, scopes=SCOPES)
    return build('drive', 'v3', credentials=creds)


def get_sheets_service():
    """Create and return Google Sheets API service."""
    creds = Credentials.from_service_account_file(SERVICE_ACCOUNT_FILE, scopes=['https://www.googleapis.com/auth/spreadsheets.readonly'])
    return build('sheets', 'v4', credentials=creds)


def list_files_in_folder(service, folder_id):
    """List all files in a folder."""
    files = []
    page_token = None
    
    while True:
        query = f"'{folder_id}' in parents and trashed=false"
        results = service.files().list(
            q=query,
            pageSize=100,
            fields="files(id, name, mimeType, shortcutDetails)",
            pageToken=page_token
        ).execute()
        
        files.extend(results.get('files', []))
        page_token = results.get('nextPageToken')
        
        if not page_token:
            break
    
    return files


def resolve_shortcut(service, file_id):
    """Resolve shortcut to get target file ID and mime type."""
    file_meta = service.files().get(
        fileId=file_id,
        fields='mimeType,shortcutDetails'
    ).execute()
    
    mime_type = file_meta.get('mimeType')
    
    # Check if this is a shortcut
    if mime_type == 'application/vnd.google-apps.shortcut':
        shortcut_details = file_meta.get('shortcutDetails', {})
        target_id = shortcut_details.get('targetId')
        target_mime = shortcut_details.get('targetMimeType')
        return target_id, target_mime
    
    return file_id, mime_type


def clean_value(value):
    """Clean illegal characters from cell values."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return value
    # Remove illegal characters that openpyxl doesn't allow
    # These are: \x00-\x08, \x0b, \x0c, \x0e-\x1f
    import re
    if isinstance(value, str):
        # Replace vertical tab and other illegal chars with space
        cleaned = re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f]', ' ', value)
        return cleaned.strip() if cleaned else None
    return value


def download_spreadsheet_values(sheets_service, spreadsheet_id, spreadsheet_name, retry_count=5, base_delay=60):
    """Download all sheets from a spreadsheet as flat values."""
    import time
    
    print(f"\n  Downloading: {spreadsheet_name}")
    
    # Generate safe filename first for check
    safe_name = re.sub(r'[<>:"/\\|?*]', '_', spreadsheet_name)
    if len(safe_name) > 100:
        safe_name = safe_name[:100]
    output_path = os.path.join(OUTPUT_FOLDER, f"{safe_name}.xlsx")
    
    # Check if file already exists
    if os.path.exists(output_path):
        print(f"    File already exists, skipping download")
        return output_path
    
    # Retry logic for rate limiting
    for attempt in range(retry_count):
        try:
            # First get list of sheets
            spreadsheet = sheets_service.spreadsheets().get(spreadsheetId=spreadsheet_id).execute()
            sheets = spreadsheet.get('sheets', [])
            break
        except Exception as e:
            if 'quota' in str(e).lower() and attempt < retry_count - 1:
                wait_time = base_delay * (attempt + 1)
                print(f"    Rate limited, waiting {wait_time}s...")
                time.sleep(wait_time)
            elif 'connection' in str(e).lower() and attempt < retry_count - 1:
                wait_time = base_delay * (attempt + 1)
                print(f"    Connection error, waiting {wait_time}s...")
                time.sleep(wait_time)
            else:
                raise
    
    print(f"    Found {len(sheets)} sheets")
    
    # Create workbook for this spreadsheet
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    for sheet in sheets:
        sheet_name = sheet['properties']['title']
        # Truncate sheet name to 31 characters (Excel limit)
        if len(sheet_name) > 31:
            sheet_name = sheet_name[:31]
        print(f"    Processing sheet: {sheet_name}")
        
        # Retry logic for rate limiting
        for attempt in range(retry_count):
            try:
                # Get values only (no formulas) using the Sheets API
                result = sheets_service.spreadsheets().values().get(
                    spreadsheetId=spreadsheet_id,
                    range=sheet_name
                ).execute()
                values = result.get('values', [])
                break
            except Exception as e:
                if 'quota' in str(e).lower() and attempt < retry_count - 1:
                    wait_time = base_delay * (attempt + 1)
                    print(f"      Rate limited, waiting {wait_time}s...")
                    time.sleep(wait_time)
                elif 'connection' in str(e).lower() and attempt < retry_count - 1:
                    wait_time = base_delay * (attempt + 1)
                    print(f"      Connection error, waiting {wait_time}s...")
                    time.sleep(wait_time)
                else:
                    raise
        
        if not values:
            print(f"      No data in sheet")
            continue
        
        # Create worksheet in workbook
        ws = wb.create_sheet(title=sheet_name)
        
        # Write values (flat, no formulas) - with cleaned illegal chars
        for row_idx, row in enumerate(values, start=1):
            for col_idx, value in enumerate(row, start=1):
                try:
                    ws.cell(row=row_idx, column=col_idx, value=clean_value(value))
                except Exception as e:
                    # If still fails, use string representation
                    ws.cell(row=row_idx, column=col_idx, value=str(value) if value else None)
        
        print(f"      Wrote {len(values)} rows")
    
    wb.save(output_path)
    print(f"    Saved to: {output_path}")
    
    return output_path


def sanitize_spreadsheet_name(name):
    """Remove project code prefix from name for cleaner filename."""
    # Remove patterns like "1LC-521UD009 - " prefix
    cleaned = re.sub(r'^[A-Z0-9]+-\d+[A-Z]?\s*-\s*', '', name)
    return cleaned


def clear_existing_files():
    """Delete all existing xlsx files in the output folder."""
    if not os.path.exists(OUTPUT_FOLDER):
        return 0
    
    deleted_count = 0
    for filename in os.listdir(OUTPUT_FOLDER):
        if filename.endswith('.xlsx'):
            file_path = os.path.join(OUTPUT_FOLDER, filename)
            try:
                os.remove(file_path)
                print(f"  Deleted: {filename}")
                deleted_count += 1
            except Exception as e:
                print(f"  Failed to delete {filename}: {e}")
    return deleted_count


def main():
    print("=" * 60)
    print("Google Drive Spreadsheet Downloader")
    print("=" * 60)
    
    # Create output folder if not exists
    os.makedirs(OUTPUT_FOLDER, exist_ok=True)
    
    # Clear existing xlsx files first
    print("\nClearing existing xlsx files in output folder...")
    deleted = clear_existing_files()
    print(f"Deleted {deleted} existing file(s)")
    
    # Initialize services
    print("\nInitializing Google Drive service...")
    drive_service = get_drive_service()
    sheets_service = get_sheets_service()
    
    # List files in folder
    print(f"\nListing files in folder: {FOLDER_ID}")
    files = list_files_in_folder(drive_service, FOLDER_ID)
    print(f"Found {len(files)} files")
    
    # Process each file
    success_count = 0
    error_count = 0
    
    for file in files:
        file_id = file['id']
        file_name = file['name']
        mime_type = file['mimeType']
        
        print(f"\nProcessing: {file_name}")
        
        try:
            # Resolve shortcut if needed
            resolved_mime = mime_type
            if mime_type == 'application/vnd.google-apps.shortcut':
                target_id, target_mime = resolve_shortcut(drive_service, file_id)
                if target_id:
                    print(f"  Shortcut resolved to: {target_id} ({target_mime})")
                    file_id = target_id
                    resolved_mime = target_mime
                else:
                    print(f"  ERROR: Could not resolve shortcut")
                    error_count += 1
                    continue
            
            # Check if it's a spreadsheet
            if resolved_mime != 'application/vnd.google-apps.spreadsheet':
                print(f"  Skipping non-spreadsheet: {resolved_mime}")
                continue
            
            # Download spreadsheet data
            output_path = download_spreadsheet_values(sheets_service, file_id, file_name)
            success_count += 1
            
        except Exception as e:
            print(f"  ERROR: {e}")
            import traceback
            traceback.print_exc()
            error_count += 1
        
        # Delay between files to avoid rate limiting
        import time
        print(f"\n  Waiting 5 seconds before next file...")
        time.sleep(5)
    
    print("\n" + "=" * 60)
    print(f"Download complete!")
    print(f"  Success: {success_count}")
    print(f"  Errors: {error_count}")
    print(f"  Output folder: {OUTPUT_FOLDER}")
    print("=" * 60)


if __name__ == '__main__':
    main()
